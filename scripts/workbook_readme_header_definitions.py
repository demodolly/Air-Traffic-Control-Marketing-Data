"""Shared README text for Process / Lead Category / Use Case / Activity Type."""

from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

HEADER_DEFINITION_LINES: list[tuple[str, bool, int | None]] = [
    ("", False, None),
    ("Use case tab headers (rows 3–6 on each UC sheet)", True, None),
    (
        "Column A lists the classification field; column B holds the value for that use case tab. "
        "Together they describe how the business classifies the touchpoint before the data element sections.",
        False,
        None,
    ),
    ("", False, None),
    ("Process", True, None),
    (
        "How the business refers to the overall action or channel family—for example Content Syndication, "
        "Pathfactory Webinars, Manual Uploads, or Online Forms.",
        False,
        None,
    ),
    ("", False, None),
    ("Lead Category", True, None),
    (
        "Determines what happens with the record after capture. Hand Raiser activities are processed "
        "for routing to the sales team. Non-Hand Raiser activities are still captured as transactions "
        "for future nurturing and lead scoring.",
        False,
        None,
    ),
    ("", False, None),
    ("Use Case", True, None),
    (
        "Distinguishes the specific type of action within the Process. This aligns with the FY27 Vehicles "
        "and Activities documentation—the content vehicle. Examples: for online forms, Offers - Contact Us "
        "or Offers - Demos; for Pathfactory Webinars, Events - Webinars.",
        False,
        None,
    ),
    ("", False, None),
    ("Activity Type", True, None),
    (
        "The most granular level of activity captured within the Use Case, also aligned to FY27 Vehicles "
        "and Activities. This is the level of detail used in attribution. Example: for Events - Webinars, "
        "activities include Registered, Attended Live, and Attended Virtual.",
        False,
        None,
    ),
]

SECTION_TITLE = "Use case tab headers (rows 3–6 on each UC sheet)"


def _last_content_row(ws: Worksheet) -> int:
    last = 1
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value not in (None, ""):
            last = row
    return last


def find_section_start(ws: Worksheet) -> int | None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == SECTION_TITLE:
            return row
    return None


def clear_header_definition_block(ws: Worksheet, start_row: int) -> None:
    for row in range(start_row, start_row + len(HEADER_DEFINITION_LINES) + 2):
        cell = ws.cell(row=row, column=1)
        cell.value = None
        cell.font = Font()


def append_header_definitions(ws: Worksheet, *, replace_existing: bool = True) -> int:
    existing = find_section_start(ws)
    if existing is not None:
        if not replace_existing:
            return existing
        clear_header_definition_block(ws, existing)
        start = existing
    else:
        start = _last_content_row(ws) + 1

    for offset, (text, bold, size) in enumerate(HEADER_DEFINITION_LINES):
        row = start + offset
        cell = ws.cell(row=row, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=size or 11)
    return start

