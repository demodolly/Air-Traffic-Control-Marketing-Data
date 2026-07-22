#!/usr/bin/env python3
"""Update Pathfactory Events - Webinar Attend Live audit from Pathfactory Attendance data."""

from __future__ import annotations

import openpyxl
from openpyxl.cell.cell import MergedCell

PATH = "/workspace/Pathfactory Events - Webinar Attend Live.xlsx"
SHEET = "Sheet1"

COL = {
    "element": 1,
    "display": 2,
    "why": 3,
    "source": 4,
    "rules": 5,
    "prompt": 6,
    "input_type": 7,
    "questions": 8,
    "attribution": 9,
    "operational": 10,
    "lead_scoring": 11,
    "seller": 12,
    "amanda": 13,
}


def set_row(ws, row: int, **kwargs) -> None:
    for key, value in kwargs.items():
        if key in COL:
            cell = ws.cell(row=row, column=COL[key])
            if isinstance(cell, MergedCell):
                continue
            cell.value = value


def unmerge_all(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))


def replace_phrases(ws) -> None:
    replacements = [
        ("Path Factory", "Pathfactory"),
        ("Telium", "Tealium"),
        ("Marketing Automation Tool", "Eloqua"),
        ("N/A - appended by Tray.io after upload", "N/A - appended by Tray.io after Pathfactory attendance is processed in Eloqua"),
        (
            "N/A - appended by Tray.io after Eloqua form submission is processed",
            "N/A - appended by Tray.io after Pathfactory attendance is processed in Eloqua",
        ),
        (
            "N/A - appended by Tray.io after Eloqua form submission is processed when available",
            "N/A - appended by Tray.io after Pathfactory attendance is processed in Eloqua when available",
        ),
        (
            "Captured on the Eloqua registration form and enriched by Contact Database",
            "Captured from the Pathfactory attendance payload in Eloqua and enriched by Contact Database",
        ),
        (
            "Captured on the Eloqua registration form when available; used to derive Job Level and Department",
            "Captured from the Pathfactory attendance payload or Contact Database; used to derive Job Level and Department when available",
        ),
        (
            "Captured on the Eloqua registration form when available; enriched by Contact Database when known.",
            "Captured from the Pathfactory attendance payload or Contact Database when known.",
        ),
        ("Captured on the Eloqua registration form", "Captured from the Pathfactory attendance payload in Eloqua"),
        (
            "Supports consent, privacy, and compliance checks for Pathfactory Eloqua registration processing.",
            "Supports consent, privacy, and compliance checks for Pathfactory attendance processing in Eloqua.",
        ),
        (
            "Connects the Pathfactory registration to the Workfront",
            "Connects the Pathfactory attendance event to the Workfront",
        ),
        (
            "Set by Tray.io on receipt of the Pathfactory registration from Eloqua.",
            "Set by Tray.io on receipt of Pathfactory attendance webhook data from Eloqua.",
        ),
        (
            "Set by Eloqua from the Pathfactory registration form configuration",
            "Derived by Tray.io from the Pathfactory attendance use case and activity type",
        ),
        ("What type of form submit is this?", "What Pathfactory attendance activity type is this?"),
        (
            "Separated the type of Form Submit for Attribution models to score separately",
            "Separates live virtual attendance from other webinar activities for attribution and lead scoring",
        ),
        (
            "Passed from Pathfactory/Tealium on the registration journey and captured by Eloqua",
            "Passed from Pathfactory on the attendance webhook payload and captured by Eloqua",
        ),
        (
            "Passed from Pathfactory/Tealium and captured by Eloqua; aligned to Workfront content where configured",
            "Passed from Pathfactory on the attendance payload and captured by Eloqua; aligned to Workfront content where configured",
        ),
        (
            "Captured by Tealium and passed through to Eloqua on the Pathfactory registration form",
            "Captured by Tealium and passed through to Eloqua with the Pathfactory attendance payload when present",
        ),
        (
            "Captured by Tealium from the visitor URL on the Pathfactory registration journey and passed to Eloqua",
            "Captured by Tealium from the visitor URL on the Pathfactory attendance journey and passed to Eloqua",
        ),
        (
            "Captured by Tealium from the registrant path to the Pathfactory webinar registration form and passed to Eloqua",
            "Captured by Tealium from the attendee journey to the Pathfactory virtual event and passed to Eloqua",
        ),
        (
            "Captured by Tealium from the registrant entry URL and passed to Eloqua",
            "Captured by Tealium from the attendee entry URL and passed to Eloqua",
        ),
        (
            "Captured by Tealium from the Pathfactory content or registration page and passed to Eloqua",
            "Captured by Tealium from the Pathfactory virtual event or content page and passed to Eloqua",
        ),
        (
            "Captured from Google Ads click tracking parameters on the registration journey and passed to Eloqua.",
            "Captured from Google Ads click tracking parameters on the attendance journey and passed to Eloqua.",
        ),
        (
            "Captured by Tealium from paid-media click tracking on the registration journey and passed to Eloqua",
            "Captured by Tealium from paid-media click tracking on the attendance journey and passed to Eloqua",
        ),
        (
            "Captured by Tealium from the previous referrer on the registration journey and passed to Eloqua",
            "Captured by Tealium from the previous referrer on the attendance journey and passed to Eloqua",
        ),
        (
            "Captured by Tealium tracking cookie on the registration journey and passed to Eloqua.",
            "Captured by Tealium tracking cookie on the attendance journey and passed to Eloqua.",
        ),
        (
            "Captured by Tealium as a first-party cookie on the registration journey and passed to Eloqua",
            "Captured by Tealium as a first-party cookie on the attendance journey and passed to Eloqua",
        ),
        (
            "Which vendor is associated with this webinar registration?",
            "Which vendor is associated with this Pathfactory virtual event attendance?",
        ),
        (
            "Confirm whether Vendor Name is always populated for Pathfactory Eloqua registrations.",
            "Confirm whether vendor metadata is always populated for Pathfactory attendance webhooks in Eloqua.",
        ),
        (
            "Connects the Pathfactory registration to the Workfront Content taxonomy",
            "Connects the Pathfactory attendance event to the Workfront Content taxonomy",
        ),
        (
            "seller enablement, and downstream follow-up context",
            "and downstream follow-up context",
        ),
        (
            "operational reporting, seller enablement, and downstream follow-up",
            "operational reporting and downstream follow-up",
        ),
    ]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                text = cell.value
                for old, new in replacements:
                    text = text.replace(old, new)
                cell.value = text


def fix_key_rows(ws) -> None:
    set_row(
        ws,
        8,
        rules="Set by Tray.io on receipt of Pathfactory attendance webhook data from Eloqua.",
        prompt="What activity type is this live virtual attendance aligned to?",
    )
    set_row(
        ws,
        9,
        source="Tray.io derived",
        rules="Derived by Tray.io from Pathfactory attendance context and use case configuration.",
    )
    set_row(
        ws,
        20,
        element="Email",
        rules="Primary matching key; captured as Visitor Email on the Pathfactory attendance webhook payload in Eloqua.",
        prompt="What is the attendee email address?",
    )
    set_row(
        ws,
        61,
        source="Pathfactory / Eloqua",
        rules="Captured from DTID on the Pathfactory attendance webhook payload when present; Tray.io may derive reporting channel and vehicle.",
        prompt="What is the Drive To ID for this attendance event?",
    )
    set_row(
        ws,
        64,
        source="Pathfactory / Eloqua",
        rules="Captured from OID on the Pathfactory attendance webhook payload when present.",
    )
    set_row(
        ws,
        76,
        rules="Captured from Pathfactory attendance or event metadata in Eloqua when present.",
    )
    set_row(
        ws,
        77,
        rules="Passed from Pathfactory on the attendance payload and captured by Eloqua; aligned to Workfront content where configured.",
        prompt="N/A - captured from Pathfactory attendance payload in Eloqua when present",
    )


def clear_seller_enablement(ws) -> None:
    for row in range(8, ws.max_row + 1):
        element = ws.cell(row=row, column=1).value
        if not element:
            continue
        if str(element).strip().startswith(tuple(f"{i} -" for i in range(1, 10))):
            continue
        cell = ws.cell(row=row, column=COL["seller"])
        if not isinstance(cell, MergedCell):
            cell.value = ""


def insert_section7(ws) -> None:
    section8_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "8 - Activity Detail":
            section8_row = row
            break
    if section8_row is None:
        return
    if section8_row > 1 and ws.cell(row=section8_row - 1, column=1).value == "Event Name":
        return
    ws.insert_rows(section8_row, 3)
    set_row(ws, section8_row, element="7 - Event Tracking Parameters")
    set_row(
        ws,
        section8_row + 1,
        element="Event ID / EID",
        display="Required",
        why="Links live attendance to the parent event for attribution and operational reporting.",
        source="Pathfactory / Eloqua",
        rules="Captured from EID on the Pathfactory attendance webhook payload in Eloqua.",
        prompt="What is the Event ID for this live virtual attendance?",
        input_type="Text",
        attribution="Yes",
        operational="Yes",
        lead_scoring="Potential",
    )
    set_row(
        ws,
        section8_row + 2,
        element="Event Name",
        display="Required",
        why="Provides human-readable virtual event context for reporting and validation.",
        source="Pathfactory / Eloqua",
        rules="Captured from Virtual Event Name on the Pathfactory attendance webhook payload.",
        prompt="What is the name of the virtual event attended?",
        input_type="Text",
        operational="Yes",
        lead_scoring="Potential",
    )


def populate_section8(ws) -> None:
    section8_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "8 - Activity Detail":
            section8_row = row
            break
    if section8_row is None:
        return
    if ws.cell(row=section8_row + 1, column=1).value == "Attendance Date":
        return

    fields = [
        {
            "element": "Attendance Date",
            "display": "Required",
            "why": "Timestamp for attribution sequencing, lead scoring recency, and operational reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Date Submitted on the Pathfactory attendance webhook payload in Eloqua.",
            "prompt": "When was the attendance event submitted?",
            "input_type": "Date/Time",
            "attribution": "Yes",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Webhook Name",
            "display": "Required",
            "why": "Identifies the Pathfactory webhook integration for troubleshooting and operational audit.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Webhook Name on the Pathfactory attendance payload.",
            "prompt": "Which Pathfactory webhook delivered this attendance event?",
            "input_type": "Text",
            "operational": "Yes",
        },
        {
            "element": "Pathfactory Event Type",
            "display": "Required",
            "why": "Classifies the Pathfactory event for attribution models and activity routing.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Event Type on the Pathfactory attendance webhook payload.",
            "prompt": "What Pathfactory event type was recorded?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Event Time",
            "display": "Required",
            "why": "Supports precise attendance timing for scoring and operational reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Event Time on the Pathfactory attendance webhook payload.",
            "prompt": "What time did the attendance event occur?",
            "input_type": "Time",
            "attribution": "Potential",
            "operational": "Yes",
        },
        {
            "element": "Event Date",
            "display": "Required",
            "why": "Supports event-day reporting and attribution alignment.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Event Date on the Pathfactory attendance webhook payload.",
            "prompt": "What date did the attendance event occur?",
            "input_type": "Date",
            "attribution": "Potential",
            "operational": "Yes",
        },
        {
            "element": "Virtual Event Experience ID",
            "display": "Required",
            "why": "Unique Pathfactory virtual event identifier for deduplication and taxonomy alignment.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Experience ID on the attendance webhook payload.",
            "prompt": "What is the Virtual Event Experience ID?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Yes",
        },
        {
            "element": "Virtual Event URL Slug",
            "display": "Optional",
            "why": "Links attendance to the published virtual event URL for operational validation.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event URL Slug on the attendance webhook payload.",
            "prompt": "What is the virtual event URL slug?",
            "input_type": "Text",
            "operational": "Potential",
        },
        {
            "element": "Virtual Event Start Time",
            "display": "Required",
            "why": "Supports scheduling alignment between registration and live attendance reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Start Time on the attendance webhook payload.",
            "prompt": "When does the virtual event start?",
            "input_type": "Date/Time",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Virtual Event External ID 1",
            "display": "Optional",
            "why": "External reference for cross-system event alignment and reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event External ID 1 when present on the attendance payload.",
            "prompt": "What is Virtual Event External ID 1?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Virtual Event External ID 2",
            "display": "Optional",
            "why": "Additional external reference for event alignment across systems.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event External ID 2 when present on the attendance payload.",
            "prompt": "What is Virtual Event External ID 2?",
            "input_type": "Text",
            "operational": "Potential",
        },
        {
            "element": "Virtual Event External ID 3",
            "display": "Optional",
            "why": "Additional external reference for event alignment across systems.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event External ID 3 when present on the attendance payload.",
            "prompt": "What is Virtual Event External ID 3?",
            "input_type": "Text",
            "operational": "Potential",
        },
        {
            "element": "Virtual Event Session Name",
            "display": "Required",
            "why": "Identifies the session attended for attribution and operational reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Session Name on the attendance webhook payload.",
            "prompt": "Which virtual event session did the attendee join?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Virtual Event Session Slug",
            "display": "Optional",
            "why": "URL-level session identifier for operational validation.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Session Slug on the attendance webhook payload.",
            "prompt": "What is the virtual event session slug?",
            "input_type": "Text",
            "operational": "Potential",
        },
        {
            "element": "Virtual Event Session Start Time",
            "display": "Required",
            "why": "Session-level scheduling for attendance validation and reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Session Start Time on the attendance webhook payload.",
            "prompt": "When does this session start?",
            "input_type": "Date/Time",
            "operational": "Yes",
        },
        {
            "element": "Virtual Event Session External ID",
            "display": "Required",
            "why": "External session identifier for deduplication and cross-system alignment.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Session External ID on the attendance webhook payload.",
            "prompt": "What is the virtual event session external ID?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Yes",
        },
        {
            "element": "Virtual Event Session Type",
            "display": "Required",
            "why": "Distinguishes live, simulive, and related session types for scoring and reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Session Type on the attendance webhook payload.",
            "prompt": "What type of virtual event session is this?",
            "input_type": "Picklist/Text",
            "attribution": "Potential",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Virtual Event Session Status",
            "display": "Required",
            "why": "Confirms session state at time of attendance for operational quality checks.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Session Status on the attendance webhook payload.",
            "prompt": "What is the session status when attendance was recorded?",
            "input_type": "Text",
            "operational": "Yes",
        },
        {
            "element": "Virtual Event Session Engagement Time (Minutes)",
            "display": "Required",
            "why": "Core engagement metric for lead scoring and content effectiveness on live attendance.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Virtual Event Session Engagement Time (Minutes) on the attendance webhook payload.",
            "prompt": "How many minutes did the attendee engage with the live session?",
            "input_type": "Number",
            "attribution": "Yes",
            "operational": "Yes",
            "lead_scoring": "Yes",
        },
        {
            "element": "Visitor Registration Status (Session)",
            "display": "Optional",
            "why": "Links attendance to session-level registration status for funnel analysis.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Visitor Registration Status (Session) on the attendance webhook payload.",
            "prompt": "What is the visitor registration status for this session?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Visitor Registration Status (Event)",
            "display": "Optional",
            "why": "Links attendance to event-level registration status for funnel analysis.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Visitor Registration Status (Event) on the attendance webhook payload.",
            "prompt": "What is the visitor registration status for the event?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Content Journey",
            "display": "Optional",
            "why": "Captures Pathfactory content journey context for attribution analysis.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Content Journey on the attendance webhook payload when present.",
            "prompt": "What content journey preceded this live attendance?",
            "input_type": "Text",
            "attribution": "Potential",
            "lead_scoring": "Potential",
        },
        {
            "element": "Supplemental Content Engagement",
            "display": "Optional",
            "why": "Indicates whether supplemental content was engaged during or around the session.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Supplemental Content Engagement on the attendance webhook payload.",
            "prompt": "Did the attendee engage with supplemental content?",
            "input_type": "Picklist/Text",
            "lead_scoring": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Supplemental Content Time Thresholds Met",
            "display": "Optional",
            "why": "Supports threshold-based scoring for supplemental content engagement.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Supplemental Content Time Thresholds Met on the attendance webhook payload.",
            "prompt": "Which supplemental content time thresholds were met?",
            "input_type": "Text",
            "lead_scoring": "Potential",
        },
        {
            "element": "Supplemental Content Engagement Time (Seconds)",
            "display": "Optional",
            "why": "Measures supplemental content depth for lead scoring beyond live session minutes.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Supplemental Content Engagement Time (Seconds) on the attendance webhook payload.",
            "prompt": "How long did the attendee engage with supplemental content (seconds)?",
            "input_type": "Number",
            "lead_scoring": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Pathfactory Session ID",
            "display": "Required",
            "why": "Session identifier from Pathfactory for deduplication and event stitching.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Session ID on the attendance webhook payload.",
            "prompt": "What is the Pathfactory session ID for this attendance?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Yes",
        },
        {
            "element": "Team",
            "display": "Optional",
            "why": "Operational routing or reporting metadata from Pathfactory when populated.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from team on the attendance webhook payload when present.",
            "prompt": "What team value is associated with this attendance event?",
            "input_type": "Text",
            "operational": "Potential",
        },
        {
            "element": "Common Campaign ID",
            "display": "Optional",
            "why": "Cross-system campaign alignment from Pathfactory attendance payload.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from CCID on the attendance webhook payload when present.",
            "prompt": "What is the Common Campaign ID (CCID)?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Conversica Flag",
            "display": "Optional",
            "why": "Indicates Conversica routing eligibility or processing from Pathfactory metadata.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from conversica on the attendance webhook payload when present.",
            "prompt": "Is this attendance flagged for Conversica processing?",
            "input_type": "Picklist/Text",
            "operational": "Potential",
        },
        {
            "element": "Pathfactory Activity Type",
            "display": "Required",
            "why": "Pathfactory activity classification for attribution and operational reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from activity_type on the attendance webhook payload.",
            "prompt": "What is the Pathfactory activity type?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Pathfactory Asset Type",
            "display": "Optional",
            "why": "Classifies the Pathfactory asset associated with attendance for content reporting.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from asset_type on the attendance webhook payload when present.",
            "prompt": "What Pathfactory asset type is associated with this attendance?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Primary Video Time Threshold Met",
            "display": "Optional",
            "why": "Threshold-based engagement signal for lead scoring on video content within the event.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Primary Video - Time Threshold Met on the attendance webhook payload.",
            "prompt": "Which primary video time thresholds were met?",
            "input_type": "Text",
            "lead_scoring": "Potential",
            "attribution": "Potential",
        },
        {
            "element": "Contact Source Original",
            "display": "Optional",
            "why": "Preserves original contact source for attribution lineage on attendance records.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from Contact Source Original on the attendance webhook payload when present.",
            "prompt": "What was the original contact source?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
    ]

    for offset, field in enumerate(fields, start=1):
        set_row(ws, section8_row + offset, **field, amanda="")


def insert_utm_id(ws) -> None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "utm_term":
            next_row = row + 1
            if ws.cell(row=next_row, column=1).value in {"utm_id", "7 - Event Tracking Parameters", "8 - Activity Detail"}:
                if ws.cell(row=next_row, column=1).value == "utm_id":
                    return
                if ws.cell(row=next_row, column=1).value in {
                    "7 - Event Tracking Parameters",
                    "8 - Activity Detail",
                }:
                    ws.insert_rows(next_row, 1)
                    set_row(
                        ws,
                        next_row,
                        element="utm_id",
                        display="Derived",
                        why="Supports campaign instance-level attribution when multiple campaigns share naming.",
                        source="Eloqua / Tealium",
                        rules="Captured by Tealium from the attendance journey URL and passed to Eloqua with the Pathfactory payload when present.",
                        prompt="N/A - captured by Tealium and passed to Eloqua when present",
                        input_type="Text",
                        attribution="Potential",
                        operational="Potential",
                    )
                return


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    ws = wb[SHEET]
    unmerge_all(ws)
    replace_phrases(ws)
    fix_key_rows(ws)
    insert_utm_id(ws)
    insert_section7(ws)
    populate_section8(ws)
    clear_seller_enablement(ws)
    wb.save(PATH)
    print(f"Updated {PATH}")


if __name__ == "__main__":
    main()
