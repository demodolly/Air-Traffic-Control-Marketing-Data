#!/usr/bin/env python3
"""
Reformat Day 20 baseline alignment into a use-case-centric workbook.

DEPRECATED: Use `Baseline Data Use Case Alignment - Day 20 V2.xlsx` as the canonical file.
Removed from repo: Day 20.xlsx and Day 20 - Reformatted.xlsx (recover from git history to rerun).
"""

from __future__ import annotations

import os
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SOURCE_PATH = "/workspace/archive/baseline-workbooks/Baseline Data Use Case Alignment - Day 20.xlsx"
DATA_ELEMENT_ORDER_PATH = "/workspace/Data Element Order.xlsx"
OUTPUT_PATH = (
    "/workspace/archive/baseline-workbooks/Baseline Data Use Case Alignment - Day 20 - Reformatted.xlsx"
)

_section_field_order: dict[str, list[str]] = {}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
SECTION_FONT = Font(bold=True)
TITLE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(bold=True, color="FFFFFF")
META_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

INDEX_HEADERS = [
    "Use Case ID",
    "Sheet Name",
    "Process",
    "Lead Category",
    "Use Case",
    "Activity Type",
    "Data Element Count",
    "Open Questions Count",
]

USE_CASE_HEADERS = [
    "Data Element",
    "Display Status",
    "Why We Need This",
    "How It Is Captured (Value Source)",
    "Data Rules",
    "Business-Friendly Prompt",
    "Input Type",
    "Outstanding Questions",
    "Attribution",
    "Operational Reporting",
    "Lead Scoring",
    "Seller Enablement",
    "Amanda Reviewed",
    "Review Status",
]

REVIEW_STATUS_COL = len(USE_CASE_HEADERS)

# Default field order for sections not defined in Data Element Order.xlsx.
DEFAULT_SECTION_FIELD_ORDER: dict[str, list[str]] = {
    "1 - System Classification": [
        "Activity Type",
        "Use Case",
        "Lead Category",
        "Sales Routing Trigger",
    ],
    "2 - Contact Identity": [
        "Email",
        "Email Status",
        "First Name",
        "Last Name",
        "Company Name",
        "Job Title",
        "Job Level",
        "Job Department",
        "Decision Maker Type",
        "Phone Status",
        "OK to Call",
        "OK to Email",
        "Region ",
        "Partner Flag",
        "Student Flag",
        "Person Party ID",
        "UUID",
        "Marketing Automation Tool Contact ID",
    ],
    "3 - Account Identity": [
        "Website URL",
        "SFDC Account ID",
        "CR Party ID",
        "ORG Party ID",
        "Address Line 1",
        "City",
        "State/County",
        "Postal/Zip Code",
        "Country",
        "Primary Phone Number",
        "Phone Extension",
        "Number of Employees",
        "Vertical Market",
        "Named Account Flag",
        "SAVM Account Manager Email",
        "SAVM Sales Coverage Code",
        "SAVM Sales Level 1",
        "SAVM Sales Level 2",
    ],
    "4 - Compliance / Permissions": [
        "Email Opt-In Permission",
        "Primary Phone Opt-In Permission",
        "Consent / Permission Status",
        "Contact Provided Permission",
        "Cross Border Storage (China Hand Raisers and Responses Only)",
        "Share Data with Cisco Affiliates and Suppliers (China Hand Raisers and Responses Only)",
        "Korea and Vietnam Combined Consent",
    ],
    "5 - Legacy Tracking Parameters": [
        "Vehicle",
        "Campaign Name",
        "Program Name",
        "Activity Name",
        "Activity Description",
        "Activity ID",
        "Offer Type",
        "Offer Name",
        "Offer Description",
        "Offer ID",
        "Drive To ID",
        "DTID Reporting Channel",
    ],
    "6 - Future Tracking Parameters": [
        "Workfront Channel ID",
        "Workfront Content ID",
        "Content Type",
        "Primary Technology",
        "Campaign",
        "Program",
        "Funnel Stage",
        "Channel Type",
        "Channel Source",
        "utm_source",
        "utm_medium",
        "utm_campaign",
        "utm_term",
        "utm_content",
        "utm_creative",
        "Vendor",
        "Source Domain",
        "Entry URL",
        "Gated Page URL",
        "Call to Action URL",
        "Google Click ID",
        "Link Click ID",
        "C_ECID",
        "C_FPID",
        "CP_GUTC",
    ],
    "7 - Event Tracking Parameters": [
        "Event ID / EID",
        "Event Name",
        "Event Start Date",
        "Event End Date",
        "Registration Date",
        "Event Attended",
        "Attendance Date",
        "Live Attendance Date",
        "Session Name",
        "Session Topic / Track",
        "Session Attendance Date",
        "Sponsored Session Name",
        "Booth Name / Location",
        "Booth Scan Date",
        "Meeting Type",
        "Meeting Owner",
        "Meeting Date",
        "CXO / Executive Indicator",
        "CXO Event Attendance Date",
        "Customer Comments",
        "Transaction Date",
    ],
    "8 - Activity Detail": [
        "Form Submit Date",
        "Registration Date",
        "Webinar Date",
        "On-Demand View Date",
        "Transaction Date",
        "Integrate Source ID",
        "Contact Source Original",
        "Sales Conversation Interest",
        "Enquiry Type",
        "Additional Enquiry Details",
        "Customer Comments",
        "Language",
        "Privacy ID",
        "Funnel Stage",
        "Marketing Automation Tool Form ID",
        "Marketing Automation Tool Email ID",
        "Marketing Automation Tool Landing Page ID",
    ],
}


def load_section_field_order_from_spreadsheet(path: str) -> dict[str, list[str]]:
    """Parse Data Element Order workbook (section headers in row 1, fields below)."""
    if not os.path.isfile(path):
        return {}

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Data Element Order" in wb.sheetnames:
        ws = wb["Data Element Order"]
    else:
        ws = wb.active

    order: dict[str, list[str]] = {}
    for col in range(1, ws.max_column + 1):
        section = ws.cell(row=1, column=col).value
        if not section or not str(section).strip():
            continue
        section_name = str(section).strip()
        if section_name in order:
            continue

        fields: list[str] = []
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None or str(value).strip() == "":
                continue
            fields.append(str(value) if str(value).endswith(" ") else str(value).strip())

        if fields:
            order[section_name] = fields

    wb.close()
    return order


def build_section_field_order() -> dict[str, list[str]]:
    merged = dict(DEFAULT_SECTION_FIELD_ORDER)
    merged.update(load_section_field_order_from_spreadsheet(DATA_ELEMENT_ORDER_PATH))
    return merged


def section_sort_key(section_name: str) -> int:
    match = re.match(r"(\d+)", str(section_name) or "")
    return int(match.group(1)) if match else 999


def field_sort_key(section_name: str, field_name: str) -> tuple[int, int, str]:
    order = _section_field_order or DEFAULT_SECTION_FIELD_ORDER
    order_list = order.get(section_name, [])
    normalized = {name: idx for idx, name in enumerate(order_list)}
    if field_name in normalized:
        return (0, normalized[field_name], field_name)
    return (1, 999, (field_name or "").lower().strip())


def sort_rows_by_section_and_field(rows: list[tuple]) -> list[tuple]:
    return sorted(
        rows,
        key=lambda r: (
            section_sort_key(r[5]),
            field_sort_key(r[5], r[6]),
        ),
    )


def ordered_sections(rows: list[tuple]) -> list[str]:
    seen: list[str] = []
    for row in sort_rows_by_section_and_field(rows):
        section = row[5]
        if section not in seen:
            seen.append(section)
    return seen


def rows_for_section(rows: list[tuple], section: str) -> list[tuple]:
    section_rows = [r for r in rows if r[5] == section]
    return sorted(section_rows, key=lambda r: field_sort_key(section, r[6]))


def style_header_row(ws, row_num: int = 1, num_cols: int | None = None) -> None:
    cols = num_cols or ws.max_column
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def auto_width(ws, max_width: int = 55) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def merge_outstanding_questions(row: tuple) -> str | None:
    parts = []
    business = row[13]
    if business:
        parts.append(str(business).strip())
    for idx, label in ((19, "Kim"), (20, "Stephanie"), (21, "Briana")):
        value = row[idx]
        if value:
            parts.append(f"{label}: {str(value).strip()}")
    if not parts:
        return None
    return " | ".join(parts)


def load_source_rows():
    wb = openpyxl.load_workbook(SOURCE_PATH, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(row)
    wb.close()
    return rows


def build_use_case_index(rows):
    combos = []
    seen = set()
    combo_counts = defaultdict(int)
    combo_questions = defaultdict(int)
    for row in rows:
        combo = (row[1], row[2], row[3], row[4])
        combo_counts[combo] += 1
        if merge_outstanding_questions(row):
            combo_questions[combo] += 1
        if combo not in seen:
            seen.add(combo)
            combos.append(combo)

    index = []
    for idx, combo in enumerate(sorted(combos), start=1):
        uc_id = f"UC{idx:02d}"
        index.append(
            {
                "id": uc_id,
                "combo": combo,
                "field_count": combo_counts[combo],
                "question_count": combo_questions[combo],
            }
        )
    return index


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", " ", name)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:31]


def assign_sheet_names(index):
    used: set[str] = set()
    for entry in index:
        combo = entry["combo"]
        base = f"{entry['id']} {combo[2]} {combo[3]}"
        name = safe_sheet_name(base)
        if name in used:
            suffix = 2
            while True:
                candidate = safe_sheet_name(f"{entry['id']} {combo[3]} {suffix}")
                if candidate not in used:
                    name = candidate
                    break
                suffix += 1
        used.add(name)
        entry["sheet_name"] = name
    return index


def apply_review_status_validation(ws, column_letter: str, max_row: int) -> None:
    if max_row < 2:
        return
    dv = DataValidation(
        type="list",
        formula1='"Confirmed,In Review,Blocked,Not Applicable"',
        allow_blank=True,
    )
    dv.error = "Choose a review status from the list."
    dv.errorTitle = "Invalid Review Status"
    ws.add_data_validation(dv)
    dv.add(f"{column_letter}2:{column_letter}{max_row}")


def write_readme_sheet(wb):
    ws = wb.active
    ws.title = "README"
    lines = [
        "Baseline Data Use Case Alignment — Reformatted Workbook",
        "",
        "Purpose",
        "One sheet per use case so you can review all required data elements, why each",
        "is needed, and outstanding questions without scrolling a single huge table.",
        "",
        "Sheets",
        "1. Use Case Index — links each Use Case ID to its dedicated sheet tab.",
        "2. UC01 through UC32 — one tab per use case (see Sheet Name column in the index).",
        "3. Open Questions — all open items across use cases.",
        "",
        "How to use",
        "- Open Use Case Index and go to the Sheet Name tab for your use case.",
        "- Each use case sheet groups fields by Section (1–8).",
        "- Fill Review Status as you confirm elements (Confirmed / In Review / Blocked).",
        "- Use Open Questions for cross-use-case follow-ups.",
        "",
        "Source file: Baseline Data Use Case Alignment - Day 20.xlsx",
        "Field order: Data Element Order.xlsx (sections 2–6); defaults for sections 1, 7, 8.",
        "Regenerate with: python3 scripts/reformat_day20_workbook.py",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif line in ("Purpose", "Sheets", "How to use"):
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 100


def write_index_sheet(wb, index):
    ws = wb.create_sheet("Use Case Index")
    ws.append(INDEX_HEADERS)
    style_header_row(ws)
    for entry in index:
        combo = entry["combo"]
        ws.append(
            [
                entry["id"],
                entry["sheet_name"],
                combo[0],
                combo[1],
                combo[2],
                combo[3],
                entry["field_count"],
                entry["question_count"],
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    auto_width(ws)


def write_use_case_sheet(wb, entry, rows):
    uc_id = entry["id"]
    combo = entry["combo"]
    ws = wb.create_sheet(entry["sheet_name"])
    ncols = len(USE_CASE_HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = ws.cell(
        row=1,
        column=1,
        value=f"{uc_id} — {combo[0]} | {combo[1]} | {combo[2]} | {combo[3]}",
    )
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = Alignment(wrap_text=True, vertical="center")

    meta_labels = [
        ("Process", combo[0]),
        ("Lead Category", combo[1]),
        ("Use Case", combo[2]),
        ("Activity Type", combo[3]),
    ]
    for i, (label, value) in enumerate(meta_labels, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = META_FILL
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=ncols)
        ws.cell(row=i, column=2, value=value)

    header_row = 6
    for col, header in enumerate(USE_CASE_HEADERS, start=1):
        ws.cell(row=header_row, column=col, value=header)
    style_header_row(ws, header_row, ncols)

    combo_rows = [r for r in rows if (r[1], r[2], r[3], r[4]) == combo]
    section_order = ordered_sections(combo_rows)

    current_row = header_row + 1
    stripe = False
    for section in section_order:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=ncols,
        )
        sec = ws.cell(row=current_row, column=1, value=section)
        sec.fill = SECTION_FILL
        sec.font = SECTION_FONT
        current_row += 1

        for r in rows_for_section(combo_rows, section):
            values = [
                r[6],
                r[7],
                r[12],
                r[9],
                r[10],
                r[8],
                r[11],
                merge_outstanding_questions(r),
                r[14],
                r[15],
                r[16],
                r[17],
                r[18],
                None,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if stripe:
                    cell.fill = ALT_FILL
            stripe = not stripe
            current_row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    status_col = get_column_letter(REVIEW_STATUS_COL)
    apply_review_status_validation(ws, status_col, current_row - 1)
    auto_width(ws, max_width=50)


def write_open_questions_sheet(wb, index, rows):
    ws = wb.create_sheet("Open Questions")
    headers = [
        "Use Case ID",
        "Sheet Name",
        "Process",
        "Use Case",
        "Activity Type",
        "Section",
        "Data Element",
        "Outstanding Questions",
        "Why We Need This",
        "Review Status",
    ]
    ws.append(headers)
    style_header_row(ws)

    id_map = {entry["combo"]: entry for entry in index}
    question_rows = [row for row in rows if merge_outstanding_questions(row)]
    question_rows.sort(
        key=lambda r: (
            id_map[(r[1], r[2], r[3], r[4])]["id"],
            section_sort_key(r[5]),
            field_sort_key(r[5], r[6]),
        )
    )
    for row in question_rows:
        combo = (row[1], row[2], row[3], row[4])
        entry = id_map[combo]
        question = merge_outstanding_questions(row)
        ws.append(
            [
                entry["id"],
                entry["sheet_name"],
                combo[0],
                combo[2],
                combo[3],
                row[5],
                row[6],
                question,
                row[12],
                None,
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    apply_review_status_validation(ws, "J", ws.max_row)
    auto_width(ws)


def main():
    global _section_field_order
    _section_field_order = build_section_field_order()

    rows = load_source_rows()
    index = assign_sheet_names(build_use_case_index(rows))

    wb = openpyxl.Workbook()
    write_readme_sheet(wb)
    write_index_sheet(wb, index)
    for entry in index:
        write_use_case_sheet(wb, entry, rows)
    write_open_questions_sheet(wb, index, rows)

    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")
    print(f"  Use case sheets: {len(index)}")
    print(f"  Total sheets: {len(wb.sheetnames)}")
    print(f"  Open question rows: {wb['Open Questions'].max_row - 1}")


if __name__ == "__main__":
    main()
