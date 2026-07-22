#!/usr/bin/env python3
"""Update Pathfactory Hand Raiser Registrations use case from Eloqua capture reference."""

from __future__ import annotations

import openpyxl

PATH = "/workspace/Pathfactory Hand Raiser Registrations.xlsx"
SHEET = "Sheet1"

COL = {
    "element": 1,
    "display": 2,
    "why": 3,
    "source": 4,
    "rules": 5,
    "prompt": 6,
    "input_type": 7,
    "questions": 8,
    "attribution": 9,
    "operational": 10,
    "lead_scoring": 11,
    "seller": 12,
    "amanda": 13,
}


def set_row(ws, row: int, **kwargs) -> None:
    for key, value in kwargs.items():
        if key in COL:
            ws.cell(row=row, column=COL[key], value=value)


def replace_phrases(ws) -> None:
    replacements = [
        ("Path Factory", "Pathfactory"),
        ("Telium", "Tealium"),
        ("Non-Hand Raisers", "Hand Raiser"),
        ("Event - Webinars", "Events - Webinars"),
        ("N/A - appended by Tray.io after upload", "N/A - appended by Tray.io after Eloqua form submission is processed"),
        (
            "N/A - appended by Tray.io after upload if available",
            "N/A - appended by Tray.io after Eloqua form submission is processed when available",
        ),
        (
            "Provided in uploaded contact file and enriched by Contact Database",
            "Captured on the Eloqua registration form and enriched by Contact Database",
        ),
        (
            "Provided in uploaded contact file when available and used to derive Job Level & Department",
            "Captured on the Eloqua registration form when available; used to derive Job Level and Department",
        ),
        (
            "Provided in uploaded contact file when available.",
            "Captured on the Eloqua registration form when available; enriched by Contact Database when known.",
        ),
        (
            "Provided in uploaded contact file.",
            "Captured on the Eloqua registration form.",
        ),
        ("Provided in Form", "Captured on the Eloqua registration form"),
        (
            "Supports consent, privacy, and compliance checks for manual-upload processing.",
            "Supports consent, privacy, and compliance checks for Pathfactory Eloqua registration processing.",
        ),
        ("Connects the upload to the Workfront", "Connects the Pathfactory registration to the Workfront"),
        (
            "Set by Tray.io on receipt of form from Marketing Automation Tool.",
            "Set by Tray.io on receipt of the Pathfactory registration from Eloqua.",
        ),
        (
            "Set by Marketing Automation Tool",
            "Set by Eloqua from the Pathfactory registration form configuration",
        ),
        (
            "Selected by Customer or defaulted by Marketing Automation Tool",
            "Selected by the registrant on the Eloqua form or defaulted by Eloqua form logic",
        ),
        (
            "Captured by Marketing Automation Tool from Offer Consumed",
            "Passed from Pathfactory/Tealium on the registration journey and captured by Eloqua",
        ),
        (
            "Captured by Marketing Automation Tool from Offer Consumed aligned to the Workfront Project that created it",
            "Passed from Pathfactory/Tealium and captured by Eloqua; aligned to Workfront content where configured",
        ),
        (
            "N/A - Appended by Telium and Marketing Automation Tool",
            "Captured by Tealium and passed through to Eloqua on the Pathfactory registration form",
        ),
        (
            "Captured by Tealium from customers driving URL and passed to Marketing Automation Tool",
            "Captured by Tealium from the visitor URL on the Pathfactory registration journey and passed to Eloqua",
        ),
        (
            "Captured by Tealium from the Customers selection of the Call to Action to access the Gated Offer and passed to Marketing Automation Tool",
            "Captured by Tealium from the registrant path to the Pathfactory webinar registration form and passed to Eloqua",
        ),
        (
            "Captured by Tealium from the Customers first entry to .com and passed to Marketing Automation Tool",
            "Captured by Tealium from the registrant entry URL and passed to Eloqua",
        ),
        (
            "Captured by Tealium from gated offer page consumed by the customer and passed to Marketing Automation Tool",
            "Captured by Tealium from the Pathfactory content or registration page and passed to Eloqua",
        ),
        (
            "Captured from Google Ads click tracking parameters and passed to Marketing Automation Tool.",
            "Captured from Google Ads click tracking parameters on the registration journey and passed to Eloqua.",
        ),
        (
            "Captured by Tealium system-generated alphanumeric code appended to an ad's destination URL when a user clicks it and passed to Marketing Automation Tool",
            "Captured by Tealium from paid-media click tracking on the registration journey and passed to Eloqua",
        ),
        (
            "Captured by Tealium from the Customers previous state to .com and passed to Marketing Automation Tool",
            "Captured by Tealium from the previous referrer on the registration journey and passed to Eloqua",
        ),
        (
            "Captured by Tealium tracking cookie and passed to Marketing Automation Tool.",
            "Captured by Tealium tracking cookie on the registration journey and passed to Eloqua.",
        ),
        (
            "Captured by Tealium as First Party Cookie and passed to Marketing Automation Tool",
            "Captured by Tealium as a first-party cookie on the registration journey and passed to Eloqua",
        ),
        (
            "Which manual upload use case is this?",
            "Which Pathfactory webinar registration use case is this?",
        ),
        (
            "Selected by uploader after use case selection.",
            "Derived from the Pathfactory Eloqua form and webinar registration context.",
        ),
        (
            "Selected by uploader at start of process.",
            "Derived from the Pathfactory Eloqua form and webinar registration configuration.",
        ),
        ("Marketing Automation Tool/Tealium", "Eloqua / Tealium"),
        ("Customer Input/Marketing Automation Tool", "Eloqua form (registrant input)"),
    ]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                text = cell.value
                for old, new in replacements:
                    text = text.replace(old, new)
                cell.value = text


def fix_classification_and_key_rows(ws) -> None:
    ws.cell(row=1, column=1, value="UC33 — Pathfactory Webinars | Hand Raiser | Events - Webinars | Registered")
    ws.cell(row=2, column=2, value="Pathfactory Webinars")
    ws.cell(row=3, column=2, value="Hand Raiser")
    ws.cell(row=4, column=2, value="Events - Webinars")
    ws.cell(row=5, column=2, value="Registered")

    set_row(
        ws,
        11,
        element="Sales Routing Trigger",
        why="Determines whether the Pathfactory registration routes to sales as a hand raiser.",
        source="Eloqua form (registrant input)",
        rules="Mapped from the Eloqua field “Do you want to speak with Cisco Sales about…” on the Pathfactory registration form.",
        prompt="Does the registrant want to be contacted immediately by our sales team?",
        input_type="Picklist/Text",
        attribution="Yes",
        operational="Yes",
        seller="Potential",
    )

    set_row(
        ws,
        38,
        rules="Captured on the Eloqua registration form (Country); Province captured where applicable.",
        prompt="Where is the registrant located?",
    )

    set_row(
        ws,
        36,
        rules="Captured on the Eloqua registration form as Province/State where applicable; enriched by Contact Database when known.",
    )

    set_row(
        ws,
        49,
        rules="Captured on the Eloqua registration form as Business Phone; enriched by Contact Database when known.",
    )


def insert_section6_missing(ws) -> None:
    """Insert Channel Type, Channel Source, Vendor before Workfront Content ID."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value in {"Channel Type", "utm_id"}:
            return

    insert_at = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Workfront Content ID":
            insert_at = row
            break
    if insert_at is None:
        return
    ws.insert_rows(insert_at, 3)
    rows = {
        insert_at: {
            "element": "Channel Type",
            "display": "Derived",
            "why": "Identifies the driving channel for attribution, operational reporting, and funnel analysis.",
            "source": "Tray.io / Workfront derived",
            "rules": "Tray.io appends from Workfront Channel ID in the Workfront Data tables.",
            "prompt": "N/A - appended by Tray.io after Eloqua form submission is processed",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        insert_at + 1: {
            "element": "Channel Source",
            "display": "Derived",
            "why": "Identifies platform/vendor source for attribution and operational reporting.",
            "source": "Tray.io / Workfront derived",
            "rules": "Tray.io appends from Workfront Channel ID in the Workfront Data tables.",
            "prompt": "N/A - appended by Tray.io after Eloqua form submission is processed",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
            "seller": "Potential",
        },
        insert_at + 2: {
            "element": "Vendor",
            "display": "Derived",
            "why": "Identifies the Pathfactory or event vendor for operational reporting and attribution context.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Vendor Name on the Eloqua registration payload when present; may be derived from event metadata.",
            "prompt": "Which vendor is associated with this webinar registration?",
            "input_type": "Text",
            "questions": "Confirm whether Vendor Name is always populated for Pathfactory Eloqua registrations.",
            "operational": "Potential",
            "seller": "Potential",
        },
    }
    for row_num, fields in rows.items():
        set_row(ws, row_num, **fields)

    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "utm_term":
            next_row = row + 1
            if ws.cell(row=next_row, column=1).value == "utm_id":
                break
            if ws.cell(row=next_row, column=1).value == "8 - Activity Detail":
                ws.insert_rows(next_row, 1)
            set_row(
                ws,
                next_row,
                element="utm_id",
                display="Derived",
                why="Supports campaign instance-level attribution when multiple campaigns share naming.",
                source="Eloqua / Tealium",
                rules="Captured by Tealium from the registration journey URL and passed to Eloqua (utm_id).",
                prompt="N/A - captured by Tealium and passed to Eloqua on the Pathfactory registration form",
                input_type="Text",
                attribution="Potential",
                operational="Potential",
            )
            break


def insert_section7(ws) -> None:
    section8_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "8 - Activity Detail":
            section8_row = row
            break
    if section8_row is None:
        return
    if ws.cell(row=section8_row - 1, column=1).value == "Event Name":
        return
    ws.insert_rows(section8_row, 3)
    set_row(ws, section8_row, element="7 - Event Tracking Parameters")
    set_row(
        ws,
        section8_row + 1,
        element="Event ID / EID",
        display="Required",
        why="Links the registration to the parent event for attribution and operational reporting.",
        source="Eloqua / Pathfactory",
        rules="Captured from Event ID on the Eloqua registration payload.",
        prompt="What is the Event ID for this webinar registration?",
        input_type="Text",
        attribution="Yes",
        operational="Yes",
        seller="Potential",
    )
    set_row(
        ws,
        section8_row + 2,
        element="Event Name",
        display="Derived",
        why="Provides human-readable event context for reporting, validation, and seller follow-up.",
        source="Eloqua / Pathfactory",
        rules="Derived from Session Name or event metadata on the Eloqua registration payload when available.",
        prompt="N/A - derived from Pathfactory/Eloqua event metadata",
        input_type="Text",
        operational="Potential",
        seller="Potential",
    )


def populate_section8(ws) -> None:
    section8_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "8 - Activity Detail":
            section8_row = row
            break
    if section8_row is None:
        return
    if ws.cell(row=section8_row + 1, column=1).value == "Registration Date":
        return

    fields = [
        {
            "element": "Registration Date",
            "display": "Required",
            "why": "Timestamp for attribution sequencing, lead scoring recency, and operational reporting.",
            "source": "Eloqua",
            "rules": "Captured from Date Submitted or FormSubmitDate on the Eloqua registration payload.",
            "prompt": "When did the registrant submit the Pathfactory registration form?",
            "input_type": "Date/Time",
            "attribution": "Yes",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Eloqua Form ID",
            "display": "Required",
            "why": "Identifies the Eloqua form used for the Pathfactory registration for audit and troubleshooting.",
            "source": "Eloqua",
            "rules": "Captured from ELQ Form ID on the Eloqua registration payload.",
            "prompt": "Which Eloqua form captured this registration?",
            "input_type": "Text",
            "operational": "Yes",
        },
        {
            "element": "Eloqua Site ID",
            "display": "Required",
            "why": "Identifies the Eloqua site/instance for routing, audit, and regional reporting.",
            "source": "Eloqua",
            "rules": "Captured from ELQ Site ID on the Eloqua registration payload.",
            "prompt": "Which Eloqua site processed this registration?",
            "input_type": "Text",
            "operational": "Potential",
        },
        {
            "element": "Eloqua Campaign ID",
            "display": "Optional",
            "why": "Connects the registration to Eloqua campaign context for attribution alignment.",
            "source": "Eloqua",
            "rules": "Captured from Eloqua Campaign ID on the registration payload when present.",
            "prompt": "What is the Eloqua Campaign ID for this registration?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Common Campaign ID",
            "display": "Optional",
            "why": "Supports cross-system campaign alignment between Eloqua, OMS, and reporting layers.",
            "source": "Eloqua / Tealium",
            "rules": "Captured from Common Campaign ID on the Eloqua registration payload when present.",
            "prompt": "What is the Common Campaign ID for this registration?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Landing Page URL",
            "display": "Required",
            "why": "Captures the registration entry page for journey and campaign attribution.",
            "source": "Eloqua / Tealium",
            "rules": "Captured from Landing Page ID/URL or ELQ Landing Page ID on the Eloqua payload.",
            "prompt": "What landing page did the registrant use before submitting the form?",
            "input_type": "URL",
            "attribution": "Potential",
            "operational": "Yes",
        },
        {
            "element": "Thank You Page URL",
            "display": "Optional",
            "why": "Confirms post-submit conversion path for operational validation.",
            "source": "Eloqua",
            "rules": "Captured from ELQ TY Landing Page URL on the Eloqua registration payload when present.",
            "prompt": "Which thank-you page was shown after registration?",
            "input_type": "URL",
            "operational": "Potential",
        },
        {
            "element": "Registrant Webinar ID",
            "display": "Required",
            "why": "Unique registrant identifier for Pathfactory webinar processing and deduplication.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Registrant Webinar ID on the Eloqua registration payload.",
            "prompt": "What is the registrant webinar ID?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Session ID",
            "display": "Required",
            "why": "Identifies the webinar session registered for; required for event-level attribution.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Session ID on the Eloqua registration payload.",
            "prompt": "Which session did the registrant sign up for?",
            "input_type": "Text",
            "attribution": "Yes",
            "operational": "Yes",
            "seller": "Potential",
        },
        {
            "element": "Session Type",
            "display": "Required",
            "why": "Distinguishes live, simulive, and on-demand sessions for reporting and scoring.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Session Type on the Eloqua registration payload.",
            "prompt": "What type of webinar session is this registration for?",
            "input_type": "Picklist/Text",
            "attribution": "Potential",
            "operational": "Yes",
            "lead_scoring": "Potential",
        },
        {
            "element": "Session Name",
            "display": "Required",
            "why": "Human-readable session context for operational reporting and seller enablement.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Session Name on the Eloqua registration payload.",
            "prompt": "What is the name of the webinar session?",
            "input_type": "Text",
            "operational": "Yes",
            "seller": "Yes",
        },
        {
            "element": "Session Description",
            "display": "Optional",
            "why": "Additional session context for seller conversations and nurture content selection.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Session Description on the Eloqua registration payload when present.",
            "prompt": "What is the session description?",
            "input_type": "Text",
            "seller": "Potential",
        },
        {
            "element": "Session Date",
            "display": "Required",
            "why": "Supports event scheduling, operational reporting, and seller follow-up timing.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Session Date on the Eloqua registration payload.",
            "prompt": "What date is the webinar session scheduled for?",
            "input_type": "Date",
            "operational": "Yes",
            "seller": "Potential",
        },
        {
            "element": "Session Time",
            "display": "Required",
            "why": "Supports event scheduling and operational reporting across regions.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Session Time on the Eloqua registration payload.",
            "prompt": "What time is the webinar session scheduled for?",
            "input_type": "Time",
            "operational": "Yes",
        },
        {
            "element": "Session Timezone",
            "display": "Required",
            "why": "Normalizes session timing for global reporting and seller routing.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Session Timezone on the Eloqua registration payload.",
            "prompt": "What timezone applies to the session time?",
            "input_type": "Text",
            "operational": "Yes",
        },
        {
            "element": "Webinar Time (EST)",
            "display": "Derived",
            "why": "Standardized webinar time for reporting and operational dashboards.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Webinar Time EST on the Eloqua registration payload or derived from session date/time.",
            "prompt": "N/A - captured or derived from Pathfactory/Eloqua session metadata",
            "input_type": "Date/Time",
            "operational": "Potential",
        },
        {
            "element": "Webinar Event URL",
            "display": "Required",
            "why": "Provides the event link for seller follow-up and registrant communications.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Webinar Event URL on the Eloqua registration payload.",
            "prompt": "What is the URL for the webinar event?",
            "input_type": "URL",
            "operational": "Yes",
            "seller": "Yes",
        },
        {
            "element": "Preferred Language",
            "display": "Recommended",
            "why": "Supports localized nurture, communications, and regional reporting.",
            "source": "Eloqua",
            "rules": "Captured from Preferred Language or Language on the Eloqua registration payload.",
            "prompt": "What is the registrant preferred language?",
            "input_type": "Picklist/Text",
            "operational": "Potential",
            "seller": "Potential",
        },
        {
            "element": "Funnel",
            "display": "Optional",
            "why": "Aligns the registration to funnel stage for scoring and routing context.",
            "source": "Eloqua / Tealium",
            "rules": "Captured from Funnel on the Eloqua registration payload when present.",
            "prompt": "What funnel stage is associated with this registration?",
            "input_type": "Text",
            "attribution": "Potential",
            "lead_scoring": "Potential",
        },
        {
            "element": "Sub Category",
            "display": "Optional",
            "why": "Provides additional content classification for attribution and reporting.",
            "source": "Eloqua",
            "rules": "Captured from Sub Category on the Eloqua registration payload when present.",
            "prompt": "What is the content sub-category for this registration?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Org Source",
            "display": "Optional",
            "why": "Tracks organizational source metadata for operational reporting.",
            "source": "Eloqua",
            "rules": "Captured from Org Source on the Eloqua registration payload when present.",
            "prompt": "What is the org source for this registration?",
            "input_type": "Text",
            "operational": "Potential",
        },
        {
            "element": "Sales Conversation Description",
            "display": "Conditional",
            "why": "Provides free-text context for sales handoff and seller preparation.",
            "source": "Eloqua form (registrant input)",
            "rules": "Captured from “Please provide a brief description of what you would like to discuss.” when provided.",
            "prompt": "What would the registrant like to discuss?",
            "input_type": "Text",
            "seller": "Yes",
        },
        {
            "element": "IRM Enquiry",
            "display": "Optional",
            "why": "Additional routing metadata from Eloqua for sales and operational workflows.",
            "source": "Eloqua",
            "rules": "Captured from IRMEnquiry2 on the Eloqua registration payload when present.",
            "prompt": "What is the IRM enquiry value for this registration?",
            "input_type": "Text",
            "operational": "Potential",
            "seller": "Potential",
        },
        {
            "element": "Pathfactory Asset Title 1",
            "display": "Optional",
            "why": "Captures content asset context consumed before registration for attribution and seller insight.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from AI_Title1 on the Eloqua registration payload when present.",
            "prompt": "What was the first Pathfactory asset title associated with this journey?",
            "input_type": "Text",
            "attribution": "Potential",
            "seller": "Potential",
        },
        {
            "element": "Pathfactory Asset Link 1",
            "display": "Optional",
            "why": "Links to the content asset consumed before registration for seller follow-up.",
            "source": "Pathfactory / Eloqua",
            "rules": "Captured from AI_Link1 on the Eloqua registration payload when present.",
            "prompt": "What was the first Pathfactory asset URL associated with this journey?",
            "input_type": "URL",
            "seller": "Potential",
        },
        {
            "element": "Content Asset ID",
            "display": "Optional",
            "why": "Identifies the Pathfactory or content asset for taxonomy and reporting alignment.",
            "source": "Eloqua / Tealium",
            "rules": "Captured from content_asset_id on the Eloqua registration payload when present.",
            "prompt": "What is the content asset ID for this registration journey?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Previous Referrer URL",
            "display": "Optional",
            "why": "Supports journey attribution analysis before the registration form.",
            "source": "Eloqua / Tealium",
            "rules": "Captured from prevReferrer on the Eloqua registration payload when present.",
            "prompt": "What was the previous referrer URL before registration?",
            "input_type": "URL",
            "attribution": "Potential",
        },
        {
            "element": "Login Flag",
            "display": "Optional",
            "why": "Indicates whether the registrant was authenticated, affecting identity matching confidence.",
            "source": "Eloqua",
            "rules": "Captured from Login Flag on the Eloqua registration payload when present.",
            "prompt": "Was the registrant logged in when they registered?",
            "input_type": "Picklist/Text",
            "operational": "Potential",
            "lead_scoring": "Potential",
        },
        {
            "element": "Asset Type",
            "display": "Optional",
            "why": "Classifies the Pathfactory asset type for content performance reporting.",
            "source": "Eloqua / Pathfactory",
            "rules": "Captured from Asset Type on the Eloqua registration payload when present.",
            "prompt": "What type of Pathfactory asset led to this registration?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Contact Source Original",
            "display": "Optional",
            "why": "Preserves original contact source for attribution and data lineage.",
            "source": "Eloqua",
            "rules": "Captured from Contact Source Original on the Eloqua registration payload when present.",
            "prompt": "What was the original contact source for this registrant?",
            "input_type": "Text",
            "attribution": "Potential",
            "operational": "Potential",
        },
        {
            "element": "Share Consent",
            "display": "Conditional",
            "why": "Supports privacy and compliance checks for Pathfactory Eloqua registration processing.",
            "source": "Eloqua form (registrant input)",
            "rules": "Captured from Share Consent on the Eloqua registration form when applicable.",
            "prompt": "Did the registrant provide share consent where required?",
            "input_type": "Picklist/Text",
            "operational": "Potential",
        },
        {
            "element": "Store Consent",
            "display": "Conditional",
            "why": "Supports privacy and compliance checks for Pathfactory Eloqua registration processing.",
            "source": "Eloqua form (registrant input)",
            "rules": "Captured from Store Consent on the Eloqua registration form when applicable.",
            "prompt": "Did the registrant provide store consent where required?",
            "input_type": "Picklist/Text",
            "operational": "Potential",
        },
    ]

    for offset, field in enumerate(fields, start=1):
        set_row(ws, section8_row + offset, **field)


def unmerge_all(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))


def clear_amanda_reviewed(ws) -> None:
    from openpyxl.cell.cell import MergedCell

    for row in range(8, ws.max_row + 1):
        element = ws.cell(row=row, column=1).value
        if not element:
            continue
        if str(element).strip().startswith(tuple(f"{i} -" for i in range(1, 10))):
            continue
        cell = ws.cell(row=row, column=COL["amanda"])
        if isinstance(cell, MergedCell):
            continue
        cell.value = ""


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    ws = wb[SHEET]
    unmerge_all(ws)
    replace_phrases(ws)
    fix_classification_and_key_rows(ws)
    insert_section6_missing(ws)
    insert_section7(ws)
    populate_section8(ws)
    clear_amanda_reviewed(ws)
    wb.save(PATH)
    print(f"Updated {PATH}")


if __name__ == "__main__":
    main()
