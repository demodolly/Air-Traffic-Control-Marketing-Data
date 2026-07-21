#!/usr/bin/env python3
"""Update Web Blogs / Brightcove video use case tab — Section 8 and video-specific copy."""

from __future__ import annotations

import openpyxl

PATH = "/workspace/Web Blogs Video Views.xlsx"
SHEET = "Sheet1"

# Columns (1-based): A=Data Element … N=Amanda Reviewed
COL = {
    "element": 1,
    "display": 2,
    "why": 3,
    "source": 4,
    "rules": 5,
    "prompt": 6,
    "input_type": 7,
    "questions": 8,
    "attribution": 9,
    "operational": 10,
    "lead_scoring": 11,
    "seller": 12,
    "amanda": 13,
}


def set_row(ws, row: int, **kwargs) -> None:
    for key, value in kwargs.items():
        if key in COL:
            ws.cell(row=row, column=COL[key], value=value)


def replace_phrases(ws) -> None:
    replacements = [
        ("N/A - appended by Tray.io after upload", "N/A - appended by Tray.io after Brightcove video view processing"),
        (
            "Provided in uploaded contact file and enriched by Contact Database",
            "Captured by Brightcove when the viewer is identified; enriched by Contact Database after matching",
        ),
        (
            "Provided in uploaded contact file when available and used to derive Job Level & Department",
            "Captured by Brightcove when available; used to derive Job Level and Department when present",
        ),
        (
            "Provided in uploaded contact file when available.",
            "Captured by Brightcove when the viewer is identified; otherwise enriched by Contact Database.",
        ),
        ("Provided in Form", "Captured by Brightcove when the viewer is known or authenticated"),
        (
            "Supports consent, privacy, and compliance checks for manual-upload processing.",
            "Supports consent, privacy, and compliance checks for video view processing and contact matching.",
        ),
        (
            "Connects the upload to the Workfront channel taxonomy",
            "Connects the video view to the Workfront channel taxonomy",
        ),
        (
            "Connects the upload to the Workfront Content taxonomy",
            "Connects the video view to the Workfront Content taxonomy",
        ),
        (
            "N/A - appended by Tray.io after upload if available",
            "N/A - appended by Tray.io after Brightcove video view processing when available",
        ),
        ("Whant is the Channel ID?", "What is the Workfront Channel ID for this video?"),
        (
            "N/A - Appended by Telium and Marketing Automation Tool",
            "N/A - derived from Brightcove video metadata and Tealium/tag management when configured",
        ),
    ]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                text = cell.value
                for old, new in replacements:
                    text = text.replace(old, new)
                cell.value = text


def fix_named_rows(ws) -> None:
  set_row(
      ws,
      3,
      element="UC33 — Brightcove Videos | Non-Hand Raiser | Web Blogs | Video View",
  )

  set_row(
      ws,
      16,
      element="Contact Database Contact ID",
      why="Provides the matched contact identifier for deduplication, routing, and auditability when the viewer is a known contact.",
      source="Tray.io / Contact Database derived",
      rules="Tray.io appends the Contact Database contact identifier after matching the Brightcove viewer; left blank for anonymous views.",
      prompt="N/A - appended by Tray.io after Brightcove video view processing",
      questions="Maps to the Contact ID row in the Data Element Matrix when this use case is added to Day 20 V2.",
  )

  set_row(
      ws,
      22,
      why="Primary contact matching key for Contact Database/SFDC matching and deduplication when the viewer is identified.",
      rules="Captured by Brightcove when the viewer is known or authenticated; may be enriched by Contact Database.",
      questions="Confirm whether business email is required for identified viewers and how anonymous views are handled.",
  )

  set_row(
      ws,
      55,
      rules="Tray.io appends campaign information from OMS tables populated by CTT Tool when aligned to the video view context",
      prompt="N/A - appended by Tray.io after Brightcove video view processing",
  )
  set_row(
      ws,
      56,
      rules="Tray.io appends program information from OMS tables populated by CTT Tool when aligned to the video view context",
      prompt="N/A - appended by Tray.io after Brightcove video view processing",
  )

  set_row(
      ws,
      64,
      why="Connects the video view to the Workfront channel taxonomy for durable operational and attribution alignment.",
      rules="Captured from Brightcove video metadata or mapping aligned to Workfront Channel ID.",
      questions="Will all Brightcove videos be set up via Workfront? Can Brightcove pass Channel ID on every view event?",
  )

  set_row(
      ws,
      71,
      why="Connects the video view to the Workfront Content taxonomy for durable operational and attribution alignment.",
      rules="Captured from Brightcove video metadata and mapped to Workfront Content ID.",
      prompt="N/A - derived from Brightcove video metadata and Tealium/tag management when configured",
  )

  section_8 = {
      74: {
          "element": "Video Name",
          "display": "Required",
          "why": "Identifies the video asset for content performance, attribution, and seller enablement reporting.",
          "source": "Brightcove",
          "rules": "Video title from Brightcove metadata passed on the view event.",
          "prompt": "What is the name of the video that was viewed?",
          "input_type": "Text",
          "questions": "Confirm canonical source if title can change in Brightcove after publish.",
          "attribution": "Yes",
          "operational": "Yes",
          "seller": "Yes",
      },
      75: {
          "element": "Video ID",
          "display": "Required",
          "why": "Unique Brightcove identifier for the video; required for deduplication, content taxonomy, and reporting.",
          "source": "Brightcove",
          "rules": "Brightcove video ID on the view event payload.",
          "prompt": "What is the Brightcove Video ID?",
          "input_type": "Text",
          "attribution": "Yes",
          "operational": "Yes",
          "lead_scoring": "Yes",
      },
      76: {
          "element": "Video Session ID",
          "display": "Optional",
          "why": "Distinguishes individual view sessions for engagement analysis and event deduplication.",
          "source": "Brightcove",
          "rules": "Session or view event identifier from Brightcove when provided.",
          "prompt": "What is the unique session or view event ID for this playback?",
          "input_type": "Text",
          "questions": "Confirm Brightcove field name for session-level ID vs. video ID.",
          "operational": "Yes",
      },
      77: {
          "element": "Video Page URL",
          "display": "Required",
          "why": "Captures where the video was viewed for journey, content, and campaign attribution.",
          "source": "Brightcove / Tealium",
          "rules": "Page URL where the player rendered, from Brightcove or tag management layer.",
          "prompt": "What is the URL of the page where the video was viewed?",
          "input_type": "URL",
          "attribution": "Yes",
          "operational": "Yes",
      },
      78: {
          "element": "% of Video Watched",
          "display": "Required",
          "why": "Measures engagement depth for lead scoring, nurture eligibility, and content effectiveness.",
          "source": "Brightcove",
          "rules": "Percentage complete from Brightcove engagement metrics on the view event.",
          "prompt": "What percentage of the video did the viewer watch?",
          "input_type": "Number",
          "attribution": "Yes",
          "operational": "Yes",
          "lead_scoring": "Yes",
      },
      79: {
          "element": "Time Video Watched",
          "display": "Required",
          "why": "Captures watch duration for engagement reporting and scoring beyond percent complete.",
          "source": "Brightcove",
          "rules": "Elapsed watch time from Brightcove engagement metrics (typically seconds).",
          "prompt": "How long did the viewer watch the video?",
          "input_type": "Number",
          "attribution": "Yes",
          "operational": "Yes",
          "lead_scoring": "Yes",
      },
      80: {
          "element": "Date Video Viewed",
          "display": "Required",
          "why": "Timestamp for attribution sequencing, scoring recency, and operational reporting.",
          "source": "Brightcove",
          "rules": "View event timestamp from Brightcove in UTC; used as primary activity date where applicable.",
          "prompt": "When did the video view occur?",
          "input_type": "Date/Time",
          "attribution": "Yes",
          "operational": "Yes",
          "lead_scoring": "Yes",
      },
      81: {
          "element": "Player ID",
          "display": "Optional",
          "why": "Identifies the Brightcove player implementation for technical troubleshooting and player-level reporting.",
          "source": "Brightcove",
          "rules": "Player ID from Brightcove metadata when multiple players exist on the site.",
          "prompt": "Which Brightcove player was used for this view?",
          "input_type": "Text",
          "operational": "Yes",
      },
      82: {
          "element": "Account ID",
          "display": "Optional",
          "why": "Brightcove account identifier tying the asset to the correct publisher/account context.",
          "source": "Brightcove",
          "rules": "Brightcove account ID from the view event or API configuration.",
          "prompt": "What is the Brightcove Account ID for this video?",
          "input_type": "Text",
          "operational": "Yes",
          "questions": "Confirm whether Account ID is needed in Snowflake or only Video ID.",
      },
  }

  for row_num, fields in section_8.items():
      set_row(ws, row_num, **fields)


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    ws = wb[SHEET]
    replace_phrases(ws)
    fix_named_rows(ws)
    wb.save(PATH)
    print(f"Updated {PATH}")


if __name__ == "__main__":
    main()
