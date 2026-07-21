#!/usr/bin/env python3
"""Add reviewer comment columns, navigation links, and README guidance to Day 20 V2."""

from __future__ import annotations

import re

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

WORKBOOK_PATH = "/workspace/Baseline Data Use Case Alignment - Day 20 V2.xlsx"

REVIEWER_HEADERS = ["Kim Comments", "Stephanie Comments", "Briana Comments"]
REVIEW_STATUS_HEADER = "Review Status"
DATA_ELEMENT_HEADER = "Data Element"
OLD_LAST_COL = 14
NEW_LAST_COL = OLD_LAST_COL + len(REVIEWER_HEADERS)

INDEX_SHEET = "Use Case Index"
MATRIX_SHEET = "Data Element Matrix"
README_SHEET = "README"


def uc_sheet_names(wb: openpyxl.Workbook) -> list[str]:
    return [name for name in wb.sheetnames if name.startswith("UC")]


def sheet_name_from_index_formula(formula_text: str) -> str | None:
    match = re.search(r"'([^']+)'!\$A\$1", formula_text or "")
    return match.group(1) if match else None


def find_header_row(ws) -> int:
    for row in range(1, 25):
        if ws.cell(row=row, column=1).value == DATA_ELEMENT_HEADER:
            if ws.cell(row=row, column=2).value == "Display Status":
                return row
    raise ValueError(f"Header row not found on sheet {ws.title}")


def expand_full_width_merges(ws, old_width: int, new_width: int, header_row: int) -> None:
    to_remerge: list[tuple[int, int]] = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_col != 1 or merged.max_col != old_width:
            continue
        if merged.min_row == merged.max_row == header_row:
            try:
                ws.unmerge_cells(merged.coord)
            except KeyError:
                pass
            continue
        to_remerge.append((merged.min_row, merged.max_row))
        try:
            ws.unmerge_cells(merged.coord)
        except KeyError:
            continue
    for min_row, max_row in to_remerge:
        ws.merge_cells(
            start_row=min_row,
            start_column=1,
            end_row=max_row,
            end_column=new_width,
        )


def widen_metadata_merges(ws, old_width: int, new_width: int, header_row: int) -> None:
    to_remerge: list[tuple[int, int, int]] = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_col != 2 or merged.max_col != old_width:
            continue
        if merged.min_row == merged.max_row == header_row:
            continue
        to_remerge.append((merged.min_row, merged.max_row, merged.min_col))
        try:
            ws.unmerge_cells(merged.coord)
        except KeyError:
            continue
    for min_row, max_row, min_col in to_remerge:
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=new_width,
        )


def apply_header_row(ws, header_row: int, saved_headers: list) -> None:
    for col, value in enumerate(saved_headers, start=1):
        ws.cell(row=header_row, column=col, value=value)
    for offset, title in enumerate(REVIEWER_HEADERS):
        ws.cell(row=header_row, column=OLD_LAST_COL + offset, value=title)
    ws.cell(row=header_row, column=NEW_LAST_COL, value=saved_headers[OLD_LAST_COL - 1])


def update_review_status_validation(ws, header_row: int, last_data_row: int) -> None:
    ws.data_validations.dataValidation.clear()
    col = get_column_letter(NEW_LAST_COL)
    dv = DataValidation(
        type="list",
        formula1='"Confirmed,In Review,Blocked,Not Applicable"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    start_row = header_row + 1
    if last_data_row >= start_row:
        dv.add(f"{col}{start_row}:{col}{last_data_row}")


def unmerge_row(ws, row_number: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row_number <= merged.max_row:
            try:
                ws.unmerge_cells(merged.coord)
            except KeyError:
                pass


def add_navigation_row(ws) -> None:
    unmerge_row(ws, 2)
    link_font = Font(color="0563C1", underline="single")

    index_cell = ws.cell(row=2, column=1, value="← Use Case Index")
    index_cell.hyperlink = f"#'{INDEX_SHEET}'!A1"
    index_cell.font = link_font

    matrix_cell = ws.cell(row=2, column=NEW_LAST_COL, value="Data Element Matrix →")
    matrix_cell.hyperlink = f"#'{MATRIX_SHEET}'!A1"
    matrix_cell.font = link_font


def insert_reviewer_columns(ws) -> int:
    header_row = find_header_row(ws)
    saved_headers = [
        ws.cell(row=header_row, column=col).value
        for col in range(1, OLD_LAST_COL + 1)
    ]

    ws.insert_cols(OLD_LAST_COL, amount=len(REVIEWER_HEADERS))
    apply_header_row(ws, header_row, saved_headers)

    expand_full_width_merges(ws, OLD_LAST_COL, NEW_LAST_COL, header_row)
    widen_metadata_merges(ws, OLD_LAST_COL, NEW_LAST_COL, header_row)
    apply_header_row(ws, header_row, saved_headers)
    return header_row


def process_use_case_sheet(ws) -> None:
    ws.insert_rows(2, 1)
    header_row = insert_reviewer_columns(ws)
    add_navigation_row(ws)
    update_review_status_validation(ws, header_row, ws.max_row)


def add_index_hyperlinks(wb: openpyxl.Workbook) -> None:
    ws = wb[INDEX_SHEET]
    link_font = Font(color="0563C1", underline="single")
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        formula = getattr(cell.value, "text", None) if cell.value else None
        target = sheet_name_from_index_formula(formula) if formula else None
        if not target:
            continue
        cell.hyperlink = f"#'{target}'!A1"
        cell.font = link_font


def write_readme(ws) -> None:
    lines = [
        ("Baseline Data Use Case Alignment — Day 20 V2", True, 14),
        ("", False, None),
        ("Purpose", True, None),
        (
            "One sheet per use case to review required data elements, why each is needed, "
            "and outstanding questions. Dynamic counts on the Use Case Index and Data Element "
            "Matrix are driven by formulas — do not change them.",
            False,
            None,
        ),
        ("", False, None),
        ("Team review rules (audit)", True, None),
        (
            "• Do NOT edit formulas, structured fields, or cells outside your comment column.",
            False,
            None,
        ),
        (
            "• Do NOT change another reviewer’s comments.",
            False,
            None,
        ),
        (
            "• Add all feedback only in the column with your name: Kim Comments, "
            "Stephanie Comments, or Briana Comments (on each use case tab, before Review Status).",
            False,
            None,
        ),
        (
            "• Use Review Status only when confirming overall status for that data element row.",
            False,
            None,
        ),
        ("", False, None),
        ("Navigation", True, None),
        ("• Use Case Index — Sheet Name links open the matching use case tab.", False, None),
        ("• Each use case tab row 2 links back to Use Case Index and Data Element Matrix.", False, None),
        ("", False, None),
        ("Sheets", True, None),
        ("• Data Element Matrix — cross-use-case field presence (formula-driven).", False, None),
        ("• Ref - Master Schema — reference definitions.", False, None),
        ("• UC01–UC32 — one tab per use case.", False, None),
        ("• Open Questions — consolidated follow-ups.", False, None),
        ("", False, None),
        ("Use case tab headers (rows 3–6 on each UC sheet)", True, None),
        (
            "Column A lists the classification field; column B holds the value for that use case tab. "
            "Together they describe how the business classifies the touchpoint before the data element sections.",
            False,
            None,
        ),
        ("", False, None),
        ("Process", True, None),
        (
            "How the business refers to the overall action or channel family—for example Content Syndication, "
            "Pathfactory Webinars, Manual Uploads, or Online Forms.",
            False,
            None,
        ),
        ("", False, None),
        ("Lead Category", True, None),
        (
            "Determines what happens with the record after capture. Hand Raiser activities are processed "
            "for routing to the sales team. Non-Hand Raiser activities are still captured as transactions "
            "for future nurturing and lead scoring.",
            False,
            None,
        ),
        ("", False, None),
        ("Use Case", True, None),
        (
            "Distinguishes the specific type of action within the Process. This aligns with the FY27 Vehicles "
            "and Activities documentation—the content vehicle. Examples: for online forms, Offers - Contact Us "
            "or Offers - Demos; for Pathfactory Webinars, Events - Webinars.",
            False,
            None,
        ),
        ("", False, None),
        ("Activity Type", True, None),
        (
            "The most granular level of activity captured within the Use Case, also aligned to FY27 Vehicles "
            "and Activities. This is the level of detail used in attribution. Example: for Events - Webinars, "
            "activities include Registered, Attended Live, and Attended Virtual.",
            False,
            None,
        ),
    ]
    for row_idx, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=size or 11)
    ws.column_dimensions["A"].width = 105


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    for name in uc_sheet_names(wb):
        process_use_case_sheet(wb[name])

    add_index_hyperlinks(wb)
    write_readme(wb[README_SHEET])

    wb.save(WORKBOOK_PATH)
    print(f"Updated {WORKBOOK_PATH}")
    print(f"  Use case sheets: {len(uc_sheet_names(wb))}")
    print(f"  Added columns: {', '.join(REVIEWER_HEADERS)}")


if __name__ == "__main__":
    main()
