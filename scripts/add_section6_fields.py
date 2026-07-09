#!/usr/bin/env python3
"""Add new Future Tracking Parameters fields to Baseline Data Use Case Alignment V2."""

import openpyxl
from copy import copy

FILE_PATH = "/workspace/Baseline Data Use Case Alignment - All Use Cases V2.xlsx"
SECTION = "6 - Future Tracking Parameters"

# Field definitions: (name, content_id_only, business_confirmation, attr, op, se)
NEW_FIELDS = [
    (
        "Content Type",
        True,
        "This field relates to Web Content Type, Webinar Type and Event Type from Workfront.",
        "Potential",
        "Potential",
        "Potential",
    ),
    (
        "Primary Technology",
        False,
        None,
        "Potential",
        "Potential",
        "Potential",
    ),
    (
        "Campaign",
        False,
        None,
        "Potential",
        "Potential",
        "Potential",
    ),
    (
        "Program",
        False,
        None,
        "Potential",
        "Potential",
        "Potential",
    ),
    (
        "Funnel Stage",
        False,
        None,
        "Potential",
        "Potential",
        None,
    ),
]

PROCESS_TEMPLATES = {
    "Content Syndication": {
        "display_status": "Derived",
        "prompt": "N/A - appended by Tray.io after upload",
        "value_source": "Tray.io / Workfront derived",
        "input_type": "System lookup",
        "reporting_purpose": None,  # set per field below
    },
    "Manual Uploads": {
        "display_status": "Derived",
        "prompt": "N/A - appended by Tray.io after upload",
        "value_source": "Tray.io / Workfront derived",
        "input_type": "System lookup",
        "reporting_purpose": None,
    },
    "Online Forms": {
        "display_status": "Derived",
        "prompt": "N/A - appended by Tray.io after upload",
        "value_source": "Tray.io / Workfront derived",
        "input_type": "System lookup",
        "reporting_purpose": None,
    },
}

FIELD_REPORTING = {
    "Content Type": "Identifies the content classification for future attribution, operational, and funnel reporting aligned to Workfront Content ID.",
    "Primary Technology": "Identifies the primary technology associated with the content for future attribution, operational, and seller enablement reporting.",
    "Campaign": "Connects the touchpoint to campaign-level attribution and marketing sourced pipeline reporting aligned to Workfront Content ID and Channel ID.",
    "Program": "Connects the touchpoint to program-level attribution and marketing sourced pipeline reporting aligned to Workfront Content ID and Channel ID.",
    "Funnel Stage": "Supports funnel-based routing, lead scoring, and operational reporting aligned to Workfront Content ID and Channel ID.",
}

FIELD_DATA_RULES = {
    "Content Type": {
        "content_id_only": "Tray.io derives Content Type from Workfront Content ID alignment.",
        "both": "Tray.io derives Content Type from Workfront Content ID alignment.",
    },
    "Primary Technology": {
        "content_id_only": "Tray.io derives Primary Technology from Workfront Content ID alignment.",
        "both": "Tray.io derives Primary Technology from Workfront Content ID and Workfront Channel ID alignment.",
    },
    "Campaign": {
        "content_id_only": "Tray.io derives Campaign from Workfront Content ID alignment.",
        "both": "Tray.io derives Campaign from Workfront Content ID and Workfront Channel ID alignment.",
    },
    "Program": {
        "content_id_only": "Tray.io derives Program from Workfront Content ID alignment.",
        "both": "Tray.io derives Program from Workfront Content ID and Workfront Channel ID alignment.",
    },
    "Funnel Stage": {
        "content_id_only": "Tray.io derives Funnel Stage from Workfront Content ID alignment.",
        "both": "Tray.io derives Funnel Stage from Workfront Content ID and Workfront Channel ID alignment.",
    },
}


def build_new_row(process, lead_category, use_case, activity_type, field_def):
    name, content_id_only, business_confirmation, attr, op, se = field_def
    template = PROCESS_TEMPLATES[process]
    alignment_key = "content_id_only" if content_id_only else "both"

    return [
        None,  # Record ID - assigned later
        process,
        lead_category,
        use_case,
        activity_type,
        SECTION,
        name,
        template["display_status"],
        template["prompt"],
        template["value_source"],
        FIELD_DATA_RULES[name][alignment_key],
        template["input_type"],
        FIELD_REPORTING[name],
        business_confirmation,
        attr,
        op,
        None,  # Lead Scoring
        se,
        None,  # Amanda Reviewed - leave blank
        None,  # Kim Comments
        None,  # Stephanie Comments
        None,  # Briana Comments
    ]


def main():
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb["Sheet1"]

    # Read header and all data rows
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    # Build new row list, inserting after each Workfront Content ID in section 6
    new_rows = []
    inserted_count = 0

    for row in rows:
        new_rows.append(row)

        is_section6_content_id = (
            row[5] == SECTION and row[6] == "Workfront Content ID"
        )
        if is_section6_content_id:
            process = row[1]
            if process not in PROCESS_TEMPLATES:
                raise ValueError(f"Unknown process type: {process}")

            for field_def in NEW_FIELDS:
                new_row = build_new_row(
                    row[1], row[2], row[3], row[4], field_def
                )
                new_rows.append(new_row)
                inserted_count += 1

    # Renumber Record IDs
    for idx, row in enumerate(new_rows, start=1):
        row[0] = idx

    # Clear existing data rows and write new ones
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_data in new_rows:
        ws.append(row_data)

    wb.save(FILE_PATH)

    print(f"Inserted {inserted_count} new rows ({inserted_count // 5} combos × 5 fields)")
    print(f"Total rows now: {len(new_rows)}")
    print(f"Saved to {FILE_PATH}")


if __name__ == "__main__":
    main()
