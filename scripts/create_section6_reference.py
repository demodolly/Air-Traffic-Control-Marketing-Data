#!/usr/bin/env python3
"""Create Section 6 Future Tracking Parameters reference spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/workspace/Section 6 Future Tracking Parameters.xlsx"
SOURCE_PATH = "/workspace/Baseline Data Use Case Alignment - All Use Cases V2.xlsx"
SECTION = "6 - Future Tracking Parameters"
NEW_FIELDS = [
    "Content Type",
    "Primary Technology",
    "Campaign",
    "Program",
    "Funnel Stage",
]

HEADERS = [
    "Record ID",
    "Process",
    "Lead Category",
    "Use Case",
    "Activity Type",
    "Section",
    "Field / Question",
    "Display Status",
    "Business-friendly prompt",
    "Value Source",
    "Data Rules",
    "Input Type",
    "Reporting / Process Purpose",
    "Business Confirmation Needed",
    "Attribution",
    "Opeational Reporting",
    "Lead Scoring",
    "Seller Enablement",
    "Amanda Reviewed",
    "Kim Comments",
    "Stephanie Comments",
    "Briana Comments",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def style_header_row(ws, row_num=1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def auto_width(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def build_field_definitions_sheet(wb):
    ws_src = openpyxl.load_workbook(SOURCE_PATH, data_only=True)["Sheet1"]
    ws = wb.active
    ws.title = "Field Definitions"

    ws.append(HEADERS)
    style_header_row(ws)

    record_id = 1
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if (
            row[5] == SECTION
            and row[6] in NEW_FIELDS
            and row[1] == "Content Syndication"
            and row[2] == "Hand Raiser"
            and row[3] == "Content Syndication"
            and row[4] == "Attended on Demand"
        ):
            ws.append([record_id] + list(row[1:]))
            record_id += 1

    for row_num in range(2, ws.max_row + 1):
        if row_num % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).fill = ALT_FILL

    auto_width(ws)
    ws.freeze_panes = "A2"


def build_use_case_scope_sheet(wb):
    ws_src = openpyxl.load_workbook(SOURCE_PATH, data_only=True)["Sheet1"]
    ws = wb.create_sheet("Use Case Scope")

    scope_headers = [
        "Record ID",
        "Process",
        "Lead Category",
        "Use Case",
        "Activity Type",
        "Fields Added",
    ]
    ws.append(scope_headers)
    style_header_row(ws)

    combos = {}
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row[5] == SECTION and row[6] in NEW_FIELDS:
            combo = (row[1], row[2], row[3], row[4])
            combos.setdefault(combo, set()).add(row[6])

    for idx, (combo, fields) in enumerate(sorted(combos.items()), start=1):
        ws.append([idx, *combo, ", ".join(sorted(fields))])

    for row_num in range(2, ws.max_row + 1):
        if row_num % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).fill = ALT_FILL

    auto_width(ws)
    ws.freeze_panes = "A2"


def build_reporting_values_sheet(wb):
    ws = wb.create_sheet("Reporting Column Values")

    reporting_headers = [
        "Field / Question",
        "Attribution",
        "Operational Reporting",
        "Lead Scoring",
        "Seller Enablement",
        "Workfront Alignment",
        "Notes",
    ]
    ws.append(reporting_headers)
    style_header_row(ws)

    rows = [
        (
            "Content Type",
            "Potential",
            "Potential",
            "",
            "Potential",
            "Workfront Content ID",
            "Relates to Web Content Type, Webinar Type and Event Type from Workfront",
        ),
        (
            "Primary Technology",
            "Potential",
            "Potential",
            "",
            "Potential",
            "Workfront Content ID + Channel ID",
            "",
        ),
        (
            "Campaign",
            "Potential",
            "Potential",
            "",
            "Potential",
            "Workfront Content ID + Channel ID",
            "",
        ),
        (
            "Program",
            "Potential",
            "Potential",
            "",
            "Potential",
            "Workfront Content ID + Channel ID",
            "",
        ),
        (
            "Funnel Stage",
            "Potential",
            "Potential",
            "",
            "",
            "Workfront Content ID + Channel ID",
            "Seller Enablement left blank, consistent with Funnel Stage in Section 8",
        ),
    ]

    for row in rows:
        ws.append(list(row))

    for row_num in range(2, ws.max_row + 1):
        if row_num % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).fill = ALT_FILL

    auto_width(ws)
    ws.freeze_panes = "A2"


def main():
    wb = openpyxl.Workbook()
    build_field_definitions_sheet(wb)
    build_use_case_scope_sheet(wb)
    build_reporting_values_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH}")
    print(f"  - Field Definitions: 5 rows")
    print(f"  - Use Case Scope: 29 rows")
    print(f"  - Reporting Column Values: 5 rows")


if __name__ == "__main__":
    main()
