#!/usr/bin/env python3
"""Remove header-definition block from Day 20 V2 README (moved to Day 21 V1)."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from workbook_readme_header_definitions import SECTION_TITLE, clear_header_definition_block

WORKBOOK_PATH = "/workspace/Baseline Data Use Case Alignment - Day 20 V2.xlsx"
README_SHEET = "README"


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb[README_SHEET]
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == SECTION_TITLE:
            clear_header_definition_block(ws, row)
            break
    wb.save(WORKBOOK_PATH)
    print(f"Removed header definitions from {WORKBOOK_PATH} README")


if __name__ == "__main__":
    main()
