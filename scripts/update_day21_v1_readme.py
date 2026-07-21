#!/usr/bin/env python3
"""Append use-case tab header definitions to the README sheet (Day 21 V1 workbook only)."""

from __future__ import annotations

import shutil
from pathlib import Path

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from workbook_readme_header_definitions import append_header_definitions

WORKBOOK_PATH = Path("/workspace/Business Data Use Case Alignment - Day 21 V1.xlsx")
README_SHEET = "README"
README_TITLE = "Business Data Use Case Alignment — Day 21 V1"
DAY20_FALLBACK = Path("/workspace/Baseline Data Use Case Alignment - Day 20 V2.xlsx")


def ensure_workbook_exists() -> None:
    if WORKBOOK_PATH.is_file():
        return
    if not DAY20_FALLBACK.is_file():
        raise FileNotFoundError(
            f"Missing {WORKBOOK_PATH} and cannot seed from {DAY20_FALLBACK}"
        )
    shutil.copy2(DAY20_FALLBACK, WORKBOOK_PATH)


def main() -> None:
    ensure_workbook_exists()
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    if README_SHEET not in wb.sheetnames:
        raise ValueError(f"{WORKBOOK_PATH} has no '{README_SHEET}' sheet")
    ws = wb[README_SHEET]
    ws.cell(row=1, column=1, value=README_TITLE)
    append_header_definitions(ws, replace_existing=True)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 105)
    wb.save(WORKBOOK_PATH)
    print(f"Updated README sheet in {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
